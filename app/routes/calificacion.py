import os
import json
import zipfile
import shutil
import tempfile
from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, jsonify, session
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from app import db
from app.models import Materia, Rubrica, Tarea, Calificacion
from app.gemini import calificar_tarea
from app.utils import extension_ok, extraer_texto

calificacion_bp = Blueprint('calificacion', __name__)


def _nombre_desde_archivo(filename):
    """Extrae el nombre del estudiante desde el nombre del archivo."""
    base = os.path.splitext(os.path.basename(filename))[0]
    base = base.replace('_', ' ').replace('-', ' ').replace('.', ' ')
    partes = [p for p in base.split() if p]
    # Si empieza con número puro (cédula/ID), quitarlo
    if partes and partes[0].isdigit():
        partes = partes[1:]
    return ' '.join(p.capitalize() for p in partes) or 'Estudiante'


# ── Ruta principal ─────────────────────────────────────────────────────────────

@calificacion_bp.route('/calificar', methods=['GET', 'POST'])
@login_required
def calificar():
    materias = Materia.query.filter_by(profesor_id=current_user.id).all()

    if request.method == 'POST':
        modo       = request.form.get('modo', 'individual')
        materia_id = request.form.get('materia_id', type=int)
        rubrica_id = request.form.get('rubrica_id', type=int)
        enunciado  = request.form.get('enunciado', '').strip()

        if not all([materia_id, rubrica_id, enunciado]):
            flash('Selecciona materia, rúbrica e ingresa el enunciado.', 'error')
            return redirect(url_for('calificacion.calificar'))

        rubrica     = Rubrica.query.get_or_404(rubrica_id)
        nota_maxima = rubrica.nota_maxima

        if modo == 'zip':
            return _procesar_zip(materia_id, rubrica_id, rubrica, enunciado, nota_maxima)
        elif modo == 'texto':
            return _procesar_texto(materia_id, rubrica_id, rubrica, enunciado, nota_maxima)
        else:
            return _procesar_individual(materia_id, rubrica_id, rubrica, enunciado, nota_maxima)

    return render_template('calificar.html', materias=materias)


# ── Modo individual ────────────────────────────────────────────────────────────

def _procesar_individual(materia_id, rubrica_id, rubrica, enunciado, nota_maxima):
    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename or not extension_ok(archivo.filename):
        flash('Sube un archivo PDF o Word válido.', 'error')
        return redirect(url_for('calificacion.calificar'))

    nombre_s = secure_filename(archivo.filename)
    ruta = os.path.join(current_app.config['UPLOAD_FOLDER'], nombre_s)
    archivo.save(ruta)
    texto = extraer_texto(ruta)

    # ✅ Pasar el nombre del archivo a Gemini para extraer el nombre del estudiante
    res = calificar_tarea(enunciado, rubrica.get_criterios(), texto, nota_maxima, archivo.filename,
                           contexto_adicional=rubrica.contexto_adicional,
                           documento_fuente=rubrica.documento_fuente_texto)

    # Si Gemini no encontró nombre, usar el nombre del archivo como respaldo
    est_nombre = res.get('nombre_estudiante', '').strip()
    if not est_nombre or est_nombre == 'Estudiante' or est_nombre == 'Error':
        est_nombre = _nombre_desde_archivo(archivo.filename)

    tarea = Tarea(materia_id=materia_id, rubrica_id=rubrica_id,
                  estudiante_nombre=est_nombre, archivo_path=ruta, enunciado=enunciado)
    db.session.add(tarea)
    db.session.flush()

    calif = Calificacion(
        tarea_id=tarea.id,
        nota_ia=res['nota'], nota_final=res['nota'],
        retroalimentacion=res.get('retroalimentacion_larga', res.get('retroalimentacion', '')),
        retroalimentacion_corta=res.get('retroalimentacion_corta', ''),
        desglose_json=json.dumps(res.get('desglose', []), ensure_ascii=False),
        prompt_usado=res.get('prompt', '')
    )
    db.session.add(calif)
    db.session.commit()

    flash(f'¡Tarea de {est_nombre} calificada correctamente!', 'success')
    return redirect(url_for('calificacion.ver_resultado', tarea_id=tarea.id))


# ── Modo texto libre ───────────────────────────────────────────────────────────

def _procesar_texto(materia_id, rubrica_id, rubrica, enunciado, nota_maxima):
    texto = request.form.get('texto_respuesta', '').strip()
    if not texto:
        flash('Pega el texto de respuesta del estudiante.', 'error')
        return redirect(url_for('calificacion.calificar'))

    # ✅ Para texto, no hay nombre de archivo, se lo pasamos vacío
    res = calificar_tarea(enunciado, rubrica.get_criterios(), texto, nota_maxima, "",
                           contexto_adicional=rubrica.contexto_adicional,
                           documento_fuente=rubrica.documento_fuente_texto)

    est_nombre = res.get('nombre_estudiante', '').strip()
    if not est_nombre or est_nombre == 'Estudiante' or est_nombre == 'Error':
        est_nombre = 'Estudiante sin nombre'

    tarea = Tarea(materia_id=materia_id, rubrica_id=rubrica_id,
                  estudiante_nombre=est_nombre, enunciado=enunciado)
    db.session.add(tarea)
    db.session.flush()

    calif = Calificacion(
        tarea_id=tarea.id,
        nota_ia=res['nota'], nota_final=res['nota'],
        retroalimentacion=res.get('retroalimentacion_larga', res.get('retroalimentacion', '')),
        retroalimentacion_corta=res.get('retroalimentacion_corta', ''),
        desglose_json=json.dumps(res.get('desglose', []), ensure_ascii=False),
        prompt_usado=res.get('prompt', '')
    )
    db.session.add(calif)
    db.session.commit()

    flash(f'¡Tarea de {est_nombre} calificada correctamente!', 'success')
    return redirect(url_for('calificacion.ver_resultado', tarea_id=tarea.id))


# ── Modo ZIP masivo ────────────────────────────────────────────────────────────

def _procesar_zip(materia_id, rubrica_id, rubrica, enunciado, nota_maxima):
    archivo_zip = request.files.get('archivo_zip')
    if not archivo_zip or not archivo_zip.filename.lower().endswith('.zip'):
        flash('Sube un archivo ZIP válido.', 'error')
        return redirect(url_for('calificacion.calificar'))

    tarea_ids = []
    errores   = []

    with tempfile.TemporaryDirectory() as tmpdir:
        zip_path = os.path.join(tmpdir, 'tareas.zip')
        archivo_zip.save(zip_path)

        try:
            with zipfile.ZipFile(zip_path, 'r') as zf:
                archivos_validos = [
                    info for info in zf.infolist()
                    if not info.is_dir()
                    and not os.path.basename(info.filename).startswith('.')
                    and '__MACOSX' not in info.filename
                    and info.filename.rsplit('.', 1)[-1].lower() in {'pdf', 'doc', 'docx'}
                ]

                if not archivos_validos:
                    flash('El ZIP no contiene archivos PDF o Word válidos.', 'error')
                    return redirect(url_for('calificacion.calificar'))

                for info in archivos_validos:
                    nombre_archivo = os.path.basename(info.filename)
                    try:
                        zf.extract(info, tmpdir)
                        ruta_extraida = os.path.join(tmpdir, info.filename)
                        texto = extraer_texto(ruta_extraida)

                        # Copiar a uploads
                        nombre_safe = secure_filename(nombre_archivo)
                        destino = os.path.join(current_app.config['UPLOAD_FOLDER'], nombre_safe)
                        shutil.copy2(ruta_extraida, destino)

                        # ✅ Pasar el nombre del archivo a Gemini para extraer el nombre del estudiante
                        res = calificar_tarea(enunciado, rubrica.get_criterios(), texto, nota_maxima, nombre_archivo,
                                               contexto_adicional=rubrica.contexto_adicional,
                           documento_fuente=rubrica.documento_fuente_texto)

                        nombre_extraido = res.get('nombre_estudiante', '').strip()
                        nombre_estudiante = nombre_extraido or _nombre_desde_archivo(info.filename)

                        tarea = Tarea(materia_id=materia_id, rubrica_id=rubrica_id,
                                      estudiante_nombre=nombre_estudiante,
                                      archivo_path=destino, enunciado=enunciado)
                        db.session.add(tarea)
                        db.session.flush()

                        calif = Calificacion(
                            tarea_id=tarea.id,
                            nota_ia=res['nota'],
                            nota_final=res['nota'],
                            retroalimentacion=res.get('retroalimentacion_larga', res.get('retroalimentacion', '')),
                            retroalimentacion_corta=res.get('retroalimentacion_corta', ''),
                            desglose_json=json.dumps(res.get('desglose', []), ensure_ascii=False),
                            prompt_usado=res.get('prompt', '')
                        )
                        db.session.add(calif)
                        db.session.commit()
                        tarea_ids.append(tarea.id)

                    except Exception as e:
                        db.session.rollback()
                        errores.append(f"{nombre_archivo}: {str(e)}")

        except zipfile.BadZipFile:
            flash('El archivo ZIP está dañado o es inválido.', 'error')
            return redirect(url_for('calificacion.calificar'))

    if not tarea_ids:
        flash('No se pudo calificar ningún archivo. Revisa los errores.', 'error')
        return redirect(url_for('calificacion.calificar'))

    session['batch_ids']     = tarea_ids
    session['batch_errores'] = errores
    return redirect(url_for('calificacion.resultado_batch'))


# ── Resultados ─────────────────────────────────────────────────────────────────

@calificacion_bp.route('/resultado/<int:tarea_id>')
@login_required
def ver_resultado(tarea_id):
    tarea = Tarea.query.get_or_404(tarea_id)
    return render_template('resultado.html', tarea=tarea, calif=tarea.calificacion)


@calificacion_bp.route('/resultado/batch')
@login_required
def resultado_batch():
    ids     = session.pop('batch_ids', [])
    errores = session.pop('batch_errores', [])
    if not ids:
        return redirect(url_for('calificacion.calificar'))
    tareas = Tarea.query.filter(Tarea.id.in_(ids)).all()
    # Ordenar según el orden original
    orden  = {tid: i for i, tid in enumerate(ids)}
    tareas = sorted(tareas, key=lambda t: orden.get(t.id, 0))
    return render_template('resultado_batch.html', tareas=tareas, errores=errores)


# ── Exportar Excel ────────────────────────────────────────────────────────────

@calificacion_bp.route('/exportar/excel')
@calificacion_bp.route('/exportar/excel/<int:materia_id>')
@login_required
def exportar_excel(materia_id=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file
    from datetime import datetime as dt

    # Obtener parámetros de filtro (desde la URL)
    materia_filtro = request.args.get('materia', type=int)
    busqueda = request.args.get('busqueda', '').strip()
    
    # Construir query con los mismos filtros que el historial
    query = (Tarea.query.join(Materia)
             .filter(Materia.profesor_id == current_user.id))
    
    # Aplicar filtros si existen
    if materia_filtro:
        query = query.filter(Tarea.materia_id == materia_filtro)
    elif materia_id:
        query = query.filter(Tarea.materia_id == materia_id)
    
    if busqueda:
        query = query.filter(
            db.or_(
                Tarea.estudiante_nombre.ilike(f'%{busqueda}%'),
                Tarea.rubrica.has(Rubrica.nombre.ilike(f'%{busqueda}%'))
            )
        )
    
    query = query.order_by(Materia.nombre, Tarea.estudiante_nombre)
    tareas = query.all()

    # Si no hay datos, mostrar mensaje
    if not tareas:
        flash('No hay datos para exportar con los filtros seleccionados.', 'error')
        return redirect(url_for('calificacion.historial'))

    # ── Crear Excel ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Calificaciones'

    # ── Estilos ──
    verde_oscuro = '1B5E20'
    verde_claro = 'E8F5E9'
    gris_header = 'F5F5F5'
    blanco = 'FFFFFF'

    header_font = Font(name='Calibri', bold=True, color=blanco, size=11)
    header_fill = PatternFill('solid', fgColor=verde_oscuro)
    subheader_font = Font(name='Calibri', bold=True, color=verde_oscuro, size=10)
    subheader_fill = PatternFill('solid', fgColor=verde_claro)
    normal_font = Font(name='Calibri', size=10)
    bold_font = Font(name='Calibri', bold=True, size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def nota_fill(nota, nota_max):
        pct = nota / nota_max if nota_max else 0
        if pct >= 0.85:
            return PatternFill('solid', fgColor='C8E6C9')
        if pct >= 0.65:
            return PatternFill('solid', fgColor='FFF9C4')
        return PatternFill('solid', fgColor='FFCDD2')

    # ── Título del reporte ──
    titulo_materia = ''
    if materia_filtro:
        materia_obj = Materia.query.get(materia_filtro)
        if materia_obj:
            titulo_materia = f' — {materia_obj.nombre}'
    elif materia_id:
        materia_obj = Materia.query.get(materia_id)
        if materia_obj:
            titulo_materia = f' — {materia_obj.nombre}'

    if busqueda:
        titulo_busqueda = f' (Búsqueda: "{busqueda}")'
    else:
        titulo_busqueda = ''

    ws.merge_cells('A1:J1')
    titulo = ws['A1']
    titulo.value = f'REPORTE DE CALIFICACIONES{titulo_materia}{titulo_busqueda} — SMARTGRADER UTM'
    titulo.font = Font(name='Calibri', bold=True, size=14, color=verde_oscuro)
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    titulo.fill = PatternFill('solid', fgColor=verde_claro)
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:J2')
    sub = ws['A2']
    sub.value = f'Profesor: {current_user.nombre}   ·   Generado: {dt.now().strftime("%d/%m/%Y %H:%M")}'
    sub.font = Font(name='Calibri', italic=True, size=10, color='757575')
    sub.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    # ── Encabezados ──
    headers = ['#', 'Estudiante', 'Tarea', 'Materia', 'Nota obtenida', 'Nota máx.', 'Nota / 10',
               'Criterios (detalle)', 'Retroalimentación general', 'Retroalimentación corta']
    widths = [5, 28, 25, 22, 14, 12, 10, 55, 50, 38]

    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[4].height = 20

    # ── Datos ──
    for row_idx, tarea in enumerate(tareas, start=5):
        nota_max = tarea.rubrica.nota_maxima if tarea.rubrica else 10.0
        nota = tarea.calificacion.nota_final if tarea.calificacion else None
        nota_10 = round(nota / nota_max * 10, 1) if nota is not None and nota_max else None
        nombre_tarea = tarea.rubrica.nombre if tarea.rubrica else '—'

        desglose = tarea.calificacion.get_desglose() if tarea.calificacion else []
        if desglose:
            criterios_texto = '\n'.join(
                f"{d.get('criterio', 'Criterio')} ({d.get('peso', 0)}%): {d.get('obtenido', 0)} pts — "
                f"{'Cumple' if d.get('cumple') else 'No cumple'} — {d.get('comentario', '').strip()}"
                for d in desglose
            )
        else:
            criterios_texto = '—'

        retro_general = (tarea.calificacion.retroalimentacion.strip()
                          if tarea.calificacion and tarea.calificacion.retroalimentacion else '—')
        retro_corta = (tarea.calificacion.retroalimentacion_corta.strip()
                       if tarea.calificacion and tarea.calificacion.retroalimentacion_corta else '—')

        row_data = [
            row_idx - 4,
            tarea.estudiante_nombre,
            nombre_tarea,
            tarea.materia.nombre,
            round(nota, 2) if nota is not None else '—',
            nota_max,
            nota_10 if nota_10 is not None else '—',
            criterios_texto,
            retro_general,
            retro_corta,
        ]

        fill_fila = PatternFill('solid', fgColor=blanco) if (row_idx % 2 == 0) else PatternFill('solid', fgColor=gris_header)

        for col_idx, valor in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center_align if col_idx in (1, 5, 6, 7) else left_align
            cell.fill = fill_fila

        if nota_10 is not None:
            ws.cell(row=row_idx, column=7).fill = nota_fill(nota, nota_max)
            ws.cell(row=row_idx, column=7).font = bold_font

        # Fila más alta si hay varios criterios o texto largo de retroalimentación, para que se lea bien
        lineas_estimadas = max(len(desglose), 1, len(retro_general) // 60)
        ws.row_dimensions[row_idx].height = max(16, min(lineas_estimadas * 26, 260))

    # ── Fila de totales ──
    if tareas:
        total_row = len(tareas) + 5
        ws.merge_cells(f'A{total_row}:D{total_row}')
        total_label = ws[f'A{total_row}']
        total_label.value = f'TOTAL: {len(tareas)} estudiante{"s" if len(tareas) != 1 else ""}'
        total_label.font = subheader_font
        total_label.fill = subheader_fill
        total_label.alignment = left_align
        total_label.border = border

        notas_10 = [round(t.calificacion.nota_final / (t.rubrica.nota_maxima or 10) * 10, 1)
                    for t in tareas if t.calificacion and t.rubrica and t.rubrica.nota_maxima]
        promedio = round(sum(notas_10) / len(notas_10), 2) if notas_10 else '—'

        # Columna F: etiqueta, columna G: valor (justo bajo "Nota / 10")
        c_label = ws.cell(row=total_row, column=6, value='Promedio /10:')
        c_label.font = subheader_font
        c_label.fill = subheader_fill
        c_label.alignment = center_align
        c_label.border = border

        c_val = ws.cell(row=total_row, column=7, value=promedio)
        c_val.font = subheader_font
        c_val.fill = subheader_fill
        c_val.alignment = center_align
        c_val.border = border

        # Resto de columnas (E y las de detalle) solo llevan el color de fondo de la fila de totales
        for col_idx in (5, 8, 9, 10):
            c = ws.cell(row=total_row, column=col_idx)
            c.fill = subheader_fill
            c.border = border

        ws.row_dimensions[total_row].height = 18

    ws.freeze_panes = 'A5'

    # ── Guardar y enviar ──
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre_archivo = 'calificaciones'
    if materia_filtro:
        materia_obj = Materia.query.get(materia_filtro)
        if materia_obj:
            nombre_archivo += f'_{materia_obj.nombre.replace(" ", "_")}'
    elif materia_id:
        materia_obj = Materia.query.get(materia_id)
        if materia_obj:
            nombre_archivo += f'_{materia_obj.nombre.replace(" ", "_")}'

    if busqueda:
        nombre_archivo += f'_busqueda_{busqueda.replace(" ", "_")}'

    nombre_archivo += f'_{dt.now().strftime("%Y%m%d_%H%M")}.xlsx'

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre_archivo
    )


# ── Historial / utilidades ─────────────────────────────────────────────────────

@calificacion_bp.route('/historial')
@login_required
def historial():
    # Obtener parámetros de filtro
    materia_id = request.args.get('materia', type=int)
    busqueda = request.args.get('busqueda', '').strip()
    
    # Query base
    query = (Tarea.query.join(Materia)
             .filter(Materia.profesor_id == current_user.id))
    
    # Aplicar filtro por materia
    if materia_id:
        query = query.filter(Tarea.materia_id == materia_id)
    
    # Aplicar búsqueda - Busca en nombre de estudiante Y en nombre de tarea (rúbrica)
    if busqueda:
        query = query.filter(
            db.or_(
                Tarea.estudiante_nombre.ilike(f'%{busqueda}%'),
                Tarea.rubrica.has(Rubrica.nombre.ilike(f'%{busqueda}%'))
            )
        )
    
    # Ordenar por fecha (más recientes primero)
    tareas = query.order_by(Tarea.subida_en.desc()).all()
    
    # Obtener todas las materias del profesor para el filtro
    materias = Materia.query.filter_by(profesor_id=current_user.id).all()
    
    return render_template('historial.html', 
                          tareas=tareas, 
                          materias=materias,
                          materia_seleccionada=materia_id,
                          busqueda=busqueda)


# ✅ API CORREGIDA - Ahora devuelve enunciado, materia_id y criterios
@calificacion_bp.route('/api/rubricas/<int:materia_id>')
@login_required
def rubricas_json(materia_id):
    try:
        rubricas = Rubrica.query.filter_by(materia_id=materia_id).all()
        result = []
        for r in rubricas:
            # Obtener criterios del JSON
            criterios_raw = r.get_criterios() if hasattr(r, 'get_criterios') else []
            
            # Asegurar formato correcto - CRUCIAL: usar 'nombre' no 'criterio'
            criterios_formateados = []
            if criterios_raw:
                for c in criterios_raw:
                    if isinstance(c, dict):
                        # Si es diccionario, asegurar que tenga 'nombre'
                        criterios_formateados.append({
                            'nombre': c.get('nombre') or c.get('criterio', 'Sin nombre'),
                            'detalle': c.get('detalle', ''),
                            'peso': c.get('peso', 0)
                        })
                    else:
                        # Si es un objeto
                        criterios_formateados.append({
                            'nombre': getattr(c, 'nombre', getattr(c, 'criterio', 'Sin nombre')),
                            'detalle': getattr(c, 'detalle', ''),
                            'peso': getattr(c, 'peso', 0)
                        })
            
            # Obtener enunciado
            enunciado = r.enunciado if hasattr(r, 'enunciado') and r.enunciado else ''
            
            result.append({
                'id': r.id,
                'nombre': r.nombre,
                'nota_maxima': r.nota_maxima,
                'enunciado': enunciado,
                'materia_id': r.materia_id,
                'criterios': criterios_formateados
            })
        return jsonify(result)
    except Exception as e:
        print(f"Error en API rubricas_json: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@calificacion_bp.route('/tarea/<int:tarea_id>/renombrar', methods=['POST'])
@login_required
def renombrar_estudiante(tarea_id):
    """Permite al profesor corregir manualmente el nombre que identificó la IA,
    ya que ningún extractor automático acierta el 100% de las veces."""
    tarea = Tarea.query.get_or_404(tarea_id)
    if tarea.materia.profesor_id != current_user.id:
        flash('No tienes permiso para editar esta tarea.', 'error')
        return redirect(url_for('calificacion.historial'))

    nuevo_nombre = request.form.get('estudiante_nombre', '').strip()
    if not nuevo_nombre:
        flash('El nombre no puede estar vacío.', 'error')
    else:
        tarea.estudiante_nombre = nuevo_nombre
        db.session.commit()
        flash('Nombre del estudiante actualizado.', 'success')

    destino = request.form.get('volver_a') or url_for('calificacion.historial')
    return redirect(destino)


@calificacion_bp.route('/eliminar_tarea/<int:tarea_id>', methods=['POST'])
@login_required
def eliminar_tarea(tarea_id):
    tarea = Tarea.query.get_or_404(tarea_id)
    if tarea.materia.profesor_id != current_user.id:
        flash('No tienes permiso para eliminar esta tarea.', 'error')
        return redirect(url_for('calificacion.historial'))

    if tarea.calificacion:
        db.session.delete(tarea.calificacion)
    db.session.delete(tarea)
    db.session.commit()

    flash(f'Tarea de {tarea.estudiante_nombre} eliminada correctamente.', 'success')
    return redirect(url_for('calificacion.historial'))


@calificacion_bp.route('/api/rubricas/nueva', methods=['POST'])
@login_required
def api_nueva_rubrica():
    try:
        # Acepta tanto JSON (sin archivo) como multipart/form-data (con documento fuente adjunto)
        if request.content_type and 'multipart/form-data' in request.content_type:
            data = request.form.to_dict()
            data['criterios'] = json.loads(request.form.get('criterios', '[]'))
            archivo_fuente = request.files.get('documento_fuente')
        else:
            data = request.get_json()
            archivo_fuente = None

        if not data:
            return jsonify({'error': 'Datos no enviados'}), 400

        materia_id = data.get('materia_id')
        nombre = data.get('nombre', '').strip()
        nota_maxima = data.get('nota_maxima', 10.0)
        enunciado = data.get('enunciado', '').strip()
        contexto_adicional = data.get('contexto_adicional', '').strip()
        criterios = data.get('criterios', [])

        # Validaciones
        if not materia_id:
            return jsonify({'error': 'Materia no seleccionada'}), 400
        if not nombre:
            return jsonify({'error': 'Nombre de la tarea es requerido'}), 400
        if not enunciado:
            return jsonify({'error': 'Enunciado de la tarea es requerido'}), 400
        if not criterios or len(criterios) == 0:
            return jsonify({'error': 'Agrega al menos un criterio'}), 400

        materia_id = int(materia_id)
        nota_maxima = float(nota_maxima) if nota_maxima else 10.0

        # Verificar que la materia existe y pertenece al profesor
        materia = Materia.query.get(materia_id)
        if not materia:
            return jsonify({'error': 'Materia no encontrada'}), 404
        if materia.profesor_id != current_user.id:
            return jsonify({'error': 'No tienes permiso para crear rúbricas en esta materia'}), 403

        # Calcular total de pesos
        total_peso = sum(c.get('peso', 0) for c in criterios)
        if total_peso != 100:
            return jsonify({'error': f'Los pesos deben sumar 100%. Actualmente suman {total_peso}%'}), 400

        # Crear la rúbrica
        rubrica = Rubrica(
            materia_id=materia_id,
            nombre=nombre,
            nota_maxima=nota_maxima,
            enunciado=enunciado,
            contexto_adicional=contexto_adicional
        )

        # Guardar criterios en formato JSON
        criterios_json = []
        for c in criterios:
            criterios_json.append({
                'nombre': c.get('nombre', '').strip(),
                'detalle': c.get('detalle', '').strip(),
                'peso': c.get('peso', 0)
            })
        rubrica.set_criterios(criterios_json)

        db.session.add(rubrica)
        db.session.flush()  # asigna rubrica.id antes de guardar el documento fuente

        if archivo_fuente and archivo_fuente.filename and extension_ok(archivo_fuente.filename):
            nombre_s = secure_filename(archivo_fuente.filename)
            ruta = os.path.join(current_app.config['UPLOAD_FOLDER'], f"rubrica{rubrica.id}_{nombre_s}")
            archivo_fuente.save(ruta)
            rubrica.documento_fuente_nombre = archivo_fuente.filename
            rubrica.documento_fuente_path = ruta
            rubrica.documento_fuente_texto = extraer_texto(ruta)

        db.session.commit()

        return jsonify({
            'id': rubrica.id,
            'nombre': rubrica.nombre,
            'nota_maxima': rubrica.nota_maxima,
            'enunciado': rubrica.enunciado,
            'documento_fuente_nombre': rubrica.documento_fuente_nombre,
            'mensaje': 'Rúbrica creada correctamente'
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"Error en api_nueva_rubrica: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500